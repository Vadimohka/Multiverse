import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def export_xlsx(records: list[dict[str, Any]], metadata: dict[str, Any] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    complex_fields = {"tables", "attachments", "official_api"}
    columns = sorted({key for row in records for key in row})
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64+min(len(columns),26))}{max(len(records)+1,1)}" if columns else "A1:A1"
    for row in records:
        values = []
        for column in columns:
            value = row.get(column)
            if column in complex_fields and isinstance(value, (dict, list)):
                value = _structured_summary(column, value)
            elif isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            values.append(value)
        ws.append(values)
    for idx, col in enumerate(columns, 1):
        max_len = max([len(str(col))] + [len(str(row.get(col, ""))) for row in records])
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = min(max_len + 2, 60)
    if metadata:
        meta = wb.create_sheet("Метаданные")
        for key, value in metadata.items():
            meta.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    _append_tables_sheet(wb, records)
    _append_attachments_sheet(wb, records)
    _append_api_sheet(wb, records)
    buffer = io.BytesIO(); wb.save(buffer)
    return buffer.getvalue()


def _structured_summary(field: str, value: object) -> str:
    if field == "tables":
        items = value if isinstance(value, list) else []
        return f"{len(items)} tables / {sum(len(item.get('rows') or []) for item in items if isinstance(item, dict))} rows"
    if field == "attachments":
        items = value if isinstance(value, list) else []
        parsed = sum(1 for item in items if isinstance(item, dict) and item.get("document"))
        return f"{len(items)} attachments / {parsed} parsed"
    if field == "official_api":
        payload = value if isinstance(value, dict) else {}
        return f"official API: {len(payload.get('records') or [])} records"
    return json.dumps(value, ensure_ascii=False)


def _base_row(row: dict[str, object]) -> dict[str, object]:
    return {key: row.get(key) for key in ("source_id", "source_name", "source_section", "title", "canonical_url", "external_id")}


def _write_dynamic_sheet(wb: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    ws = wb.create_sheet(title[:31])
    columns = sorted({key for row in rows for key in row})
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([json.dumps(row.get(column), ensure_ascii=False) if isinstance(row.get(column), (dict, list)) else row.get(column) for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + min(len(columns), 26))}{max(len(rows) + 1, 1)}"
    for index, column in enumerate(columns, 1):
        width = max([len(str(column))] + [len(str(row.get(column, ""))) for row in rows])
        ws.column_dimensions[ws.cell(1, index).column_letter].width = min(width + 2, 60)


def _append_tables_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    output: list[dict[str, object]] = []
    for record in records:
        for table in record.get("tables") or []:
            if not isinstance(table, dict):
                continue
            for index, table_row in enumerate(table.get("rows") or []):
                if isinstance(table_row, dict):
                    output.append({**_base_row(record), "table_index": table.get("table_index"), "row_index": index, **table_row})
    _write_dynamic_sheet(wb, "Таблицы", output)


def _append_attachments_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    output: list[dict[str, object]] = []
    for record in records:
        for attachment in record.get("attachments") or []:
            if not isinstance(attachment, dict):
                continue
            document = attachment.get("document") if isinstance(attachment.get("document"), dict) else {}
            for page in document.get("pages") or []:
                if isinstance(page, dict) and page.get("text"):
                    output.append({
                        **_base_row(record), "attachment_title": attachment.get("title"),
                        "attachment_url": attachment.get("url"), "filename": document.get("filename"),
                        "document_page": page.get("page"), "document_text": page.get("text"),
                    })
            for sheet in document.get("sheets") or []:
                if not isinstance(sheet, dict):
                    continue
                for row_index, cells in enumerate(sheet.get("rows") or []):
                    item = {**_base_row(record), "attachment_title": attachment.get("title"), "attachment_url": attachment.get("url"), "filename": document.get("filename"), "sheet": sheet.get("name"), "document_row_index": row_index}
                    item.update({f"cell_{index + 1}": value for index, value in enumerate(cells or [])})
                    output.append(item)
    _write_dynamic_sheet(wb, "Вложения", output)


def _append_api_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    output: list[dict[str, object]] = []
    for record in records:
        payload = record.get("official_api") if isinstance(record.get("official_api"), dict) else {}
        for item in payload.get("records") or []:
            if isinstance(item, dict):
                output.append({**_base_row(record), **item})
    _write_dynamic_sheet(wb, "Официальный API", output)
